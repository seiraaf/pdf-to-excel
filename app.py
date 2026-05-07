from io import BytesIO
import re

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SUPPLIER_DB_PATH = r"C:\Users\TRM\Downloads\DATABASE SUP.xlsx"
UNITS = ["Nirwana", "Lovina", "Lembongan", "The Club", "Kanaka"]
COLUMNS = ["Unit", "Supplier", "No PO", "Item", "Qty", "Unit Item", "Price", "Amount", "Remarks"]
EXPORT_HEADERS = ["Supplier", "No PO", "Inv Name", "Q", "Unit", "Description", "Qty Datang", "Remark"]
UNIT_TITLES = {
    "Nirwana": "NIRWANA BEACH RESORT",
    "Lovina": "LOVINA BEACH RESORT",
    "Lembongan": "LEMBONGAN",
    "The Club": "THE CLUB",
    "Kanaka": "KANAKA",
}
MONTHS_ID = {
    "jan": "JANUARI",
    "january": "JANUARI",
    "feb": "FEBRUARI",
    "february": "FEBRUARI",
    "mar": "MARET",
    "march": "MARET",
    "apr": "APRIL",
    "april": "APRIL",
    "may": "MEI",
    "jun": "JUNI",
    "june": "JUNI",
    "jul": "JULI",
    "july": "JULI",
    "aug": "AGUSTUS",
    "august": "AGUSTUS",
    "sep": "SEPTEMBER",
    "sept": "SEPTEMBER",
    "september": "SEPTEMBER",
    "oct": "OKTOBER",
    "october": "OKTOBER",
    "nov": "NOVEMBER",
    "november": "NOVEMBER",
    "dec": "DESEMBER",
    "december": "DESEMBER",
}

NUMBER_PATTERN = r"\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?"
PATTERN_WITH_NO = re.compile(
    rf"^\s*\d+\s+(?P<item>.*?)\s+(?P<qty>{NUMBER_PATTERN})\s+(?P<unit>[A-Za-z]+)\s+"
    rf"(?P<price>{NUMBER_PATTERN})\s+(?P<amount>{NUMBER_PATTERN})(?:\s+(?P<tail>.*))?\s*$",
    re.IGNORECASE,
)
PATTERN_WITHOUT_NO = re.compile(
    rf"^\s*(?P<item>.*?)\s+(?P<qty>{NUMBER_PATTERN})\s+(?P<unit>[A-Za-z]+)\s+"
    rf"(?P<price>{NUMBER_PATTERN})\s+(?P<amount>{NUMBER_PATTERN})(?:\s+(?P<tail>.*))?\s*$",
    re.IGNORECASE,
)

SKIP_KEYWORDS = (
    "subtotal",
    "grand total",
    "purchase order",
    "prepared by",
    "approved by",
    "discount",
    "tax & freight",
    "total :",
)
TABLE_SETTINGS = {
    "vertical_strategy": "text",
    "horizontal_strategy": "text",
    "snap_tolerance": 4,
    "join_tolerance": 4,
    "intersection_tolerance": 6,
    "text_tolerance": 3,
}


st.set_page_config(page_title="PDF to Excel", page_icon="chart", layout="wide")


if "all_data" not in st.session_state:
    st.session_state.all_data = []

if "excel_file" not in st.session_state:
    st.session_state.excel_file = None


def clean_text(value, default="-"):
    value = re.sub(r"\s+", " ", str(value or "")).strip()
    return value if value else default


def normalize_number(value):
    return clean_text(value).replace(" ", "")


def to_excel_number(value):
    value = clean_text(value, "").replace(",", "")
    try:
        number = float(value)
    except ValueError:
        return clean_text(value)
    return int(number) if number.is_integer() else number


def normalize_lookup(value):
    value = clean_text(value, "").upper()
    value = value.replace("&", " AND ")
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return clean_text(value, "")


def compact_lookup(value):
    return normalize_lookup(value).replace(" ", "")


@st.cache_data(show_spinner=False)
def prepare_supplier_database(file_bytes=None, path=None):
    try:
        if file_bytes:
            df = pd.read_excel(BytesIO(file_bytes))
        else:
            df = pd.read_excel(path)
    except Exception:
        return pd.DataFrame(columns=["NAMA", "VENDOR", "LOOKUP", "LOOKUP_COMPACT"])

    df.columns = [clean_text(column, "").upper() for column in df.columns]
    if "NAMA" not in df.columns or "VENDOR" not in df.columns:
        return pd.DataFrame(columns=["NAMA", "VENDOR", "LOOKUP", "LOOKUP_COMPACT"])

    df = df[["NAMA", "VENDOR"]].dropna(how="any")
    df["NAMA"] = df["NAMA"].map(lambda value: clean_text(value, ""))
    df["VENDOR"] = df["VENDOR"].map(lambda value: clean_text(value, ""))
    df["LOOKUP"] = df["NAMA"].map(normalize_lookup)
    df["LOOKUP_COMPACT"] = df["NAMA"].map(compact_lookup)
    df = df[df["LOOKUP"] != ""]
    return df


def lookup_supplier(item, supplier_db):
    item_lookup = normalize_lookup(item)
    item_compact = compact_lookup(item)
    if not item_lookup or supplier_db.empty:
        return ""

    exact = supplier_db[supplier_db["LOOKUP"] == item_lookup]
    if not exact.empty:
        return exact.iloc[0]["VENDOR"]

    compact_exact = supplier_db[supplier_db["LOOKUP_COMPACT"] == item_compact]
    if not compact_exact.empty:
        return compact_exact.iloc[0]["VENDOR"]

    candidates = []
    for row in supplier_db.itertuples(index=False):
        db_item = row.LOOKUP
        db_compact = row.LOOKUP_COMPACT
        if len(db_compact) < 4:
            continue
        if db_item in item_lookup or item_lookup in db_item:
            candidates.append((len(db_item), row.VENDOR))
        elif db_compact in item_compact or item_compact in db_compact:
            candidates.append((len(db_compact), row.VENDOR))

    if candidates:
        return sorted(candidates, reverse=True)[0][1]
    return ""


def looks_like_number(value):
    return bool(re.fullmatch(NUMBER_PATTERN, clean_text(value, "")))


def looks_like_unit(value):
    value = clean_text(value, "")
    return bool(re.fullmatch(r"[A-Za-z]{1,8}", value))


def should_skip_line(line):
    lowered = line.lower()
    return not line or any(keyword in lowered for keyword in SKIP_KEYWORDS)


def format_po_date(raw_date):
    raw_date = clean_text(raw_date, "")
    match = re.search(r"(\d{1,2})[-/\s]([A-Za-z]+)[-/\s](\d{2,4})", raw_date)
    if not match:
        return ""

    day, month, year = match.groups()
    if len(year) == 2:
        year = "20" + year

    month_name = MONTHS_ID.get(month.lower(), month.upper())
    return f"PO {int(day)} {month_name} {year}"


def extract_po_date(text):
    for line in (text or "").splitlines():
        match = re.search(r"\b(?:Date|Tanggal)\b\s*:?[\s-]*(.+)$", line, re.IGNORECASE)
        if match:
            po_date = format_po_date(match.group(1))
            if po_date:
                return po_date
    return ""


def extract_no_po(text):
    match = re.search(r"\bPO/\d+/\d+\b", text or "", re.IGNORECASE)
    return clean_text(match.group(0), "") if match else ""


def extract_header_remark(text):
    for line in (text or "").splitlines():
        match = re.search(r"\bRemark\b\s*:?[\s-]*(.+)$", line, re.IGNORECASE)
        if match:
            remark = clean_text(match.group(1), "")
            if remark and not remark.lower().startswith(("price", "amount", "supplier")):
                return remark
    return "-"


def extract_row_remark(tail, header_remark="-"):
    tail = clean_text(tail, "")
    header_remark = clean_text(header_remark)
    if not tail:
        return header_remark

    # Format PO kedua: Amount Remark PR No.
    tail_without_pr = re.split(r"\bPR/\d+/\d+\b", tail, maxsplit=1, flags=re.IGNORECASE)[0]
    tail_without_pr = clean_text(tail_without_pr, "")
    if tail_without_pr:
        return tail_without_pr

    # Format Daily Market List biasanya supplier ada setelah Amount, remarks ada di header.
    return header_remark


def make_row(unit, item, qty, unit_item, price, amount, remarks, no_po, supplier_db):
    return {
        "Unit": unit,
        "Supplier": lookup_supplier(item, supplier_db),
        "No PO": no_po,
        "Item": clean_text(item),
        "Qty": normalize_number(qty),
        "Unit Item": clean_text(unit_item).upper(),
        "Price": normalize_number(price),
        "Amount": normalize_number(amount),
        "Remarks": clean_text(remarks),
    }


def parse_line(line, unit, header_remark="-", po_date="", no_po="", supplier_db=None):
    line = clean_text(line, "")
    if should_skip_line(line):
        return None

    match = PATTERN_WITH_NO.match(line) or PATTERN_WITHOUT_NO.match(line)
    if not match:
        return None

    data = match.groupdict()
    remarks = extract_row_remark(data.get("tail", ""), header_remark)
    row = make_row(
        unit,
        data["item"],
        data["qty"],
        data["unit"],
        data["price"],
        data["amount"],
        remarks,
        no_po,
        supplier_db,
    )
    row["PO Date"] = po_date
    return row


def parse_table_row(cells, unit, header_remark="-", po_date="", no_po="", supplier_db=None):
    cells = [clean_text(cell, "") for cell in cells]
    cells = [cell for cell in cells if cell]
    line = " ".join(cells)

    if should_skip_line(line) or "item name" in line.lower() or len(cells) < 4:
        return None

    number_indexes = [index for index, cell in enumerate(cells) if looks_like_number(cell)]
    if len(number_indexes) < 3:
        return None

    qty_index = number_indexes[0]
    amount_index = number_indexes[-1]
    price_index = number_indexes[-2]

    unit_index = None
    for index in range(qty_index + 1, price_index):
        if looks_like_unit(cells[index]):
            unit_index = index
            break

    if unit_index is None:
        return None

    item = clean_text(" ".join(cells[:qty_index]), "")
    if not item:
        return None

    tail = clean_text(" ".join(cells[amount_index + 1 :]), "")
    remarks = extract_row_remark(tail, header_remark)
    row = make_row(
        unit,
        item,
        cells[qty_index],
        cells[unit_index],
        cells[price_index],
        cells[amount_index],
        remarks,
        no_po,
        supplier_db,
    )
    row["PO Date"] = po_date
    return row


def row_key(row):
    return tuple(row[column] for column in ["Unit", "Item", "Qty", "Unit Item", "Price", "Amount"])


def parse_pdf(uploaded_file, unit, supplier_db):
    rows = []
    seen = set()

    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            header_remark = extract_header_remark(text)
            po_date = extract_po_date(text)
            no_po = extract_no_po(text)

            for table in page.extract_tables(TABLE_SETTINGS) or []:
                for cells in table:
                    row = parse_table_row(cells, unit, header_remark, po_date, no_po, supplier_db)
                    if row and row_key(row) not in seen:
                        rows.append(row)
                        seen.add(row_key(row))

            for line in text.splitlines():
                row = parse_line(line, unit, header_remark, po_date, no_po, supplier_db)
                if row and row_key(row) not in seen:
                    rows.append(row)
                    seen.add(row_key(row))

    return rows


def export_row(row):
    return [
        row.get("Supplier", ""),
        row.get("No PO", ""),
        row.get("Item", ""),
        to_excel_number(row.get("Qty", "")),
        row.get("Unit Item", ""),
        "",
        "",
        "" if row.get("Remarks") == "-" else row.get("Remarks", ""),
    ]


def style_sheet(ws, unit_name, po_date, last_row):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws["A1"] = UNIT_TITLES.get(unit_name, unit_name.upper())
    ws["A2"] = po_date or "PO"
    ws["A1"].font = Font(bold=True, size=20)
    ws["A2"].font = Font(bold=True, size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 18

    widths = [18, 18, 44, 10, 8, 48, 16, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for cell in ws[4]:
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=5, max_row=max(last_row, 5), min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(size=10)
        row[2].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row[3].alignment = Alignment(horizontal="right", vertical="center")
        row[4].alignment = Alignment(horizontal="center", vertical="center")
        row[3].number_format = "#,##0.00"

    ws.auto_filter.ref = f"A4:H{max(last_row, 5)}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = True


def build_excel(data):
    output = BytesIO()
    df = pd.DataFrame(data)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for unit_name in UNITS:
        ws = workbook.create_sheet(unit_name[:31])
        unit_df = df[df["Unit"] == unit_name] if not df.empty and "Unit" in df.columns else pd.DataFrame()
        po_dates = []
        if not unit_df.empty and "PO Date" in unit_df.columns:
            po_dates = [clean_text(value, "") for value in unit_df["PO Date"] if clean_text(value, "")]
        po_date = po_dates[0] if po_dates else ""

        ws.append([""] * len(EXPORT_HEADERS))
        ws.append([""] * len(EXPORT_HEADERS))
        ws.append([""] * len(EXPORT_HEADERS))
        ws.append(EXPORT_HEADERS)

        if not unit_df.empty:
            for row in unit_df.to_dict("records"):
                ws.append(export_row(row))

        style_sheet(ws, unit_name, po_date, ws.max_row)

    workbook.save(output)
    output.seek(0)
    return output


st.title("PDF to Excel Converter")
st.caption("Upload per unit -> Tambah Data -> Convert Semua Unit -> download 1 Excel multi sheet.")

with st.sidebar:
    st.header("Database Supplier")
    supplier_file = st.file_uploader("Upload DATABASE SUP.xlsx", type=["xlsx"], key="supplier_db")
    if supplier_file:
        supplier_db = prepare_supplier_database(supplier_file.getvalue())
    else:
        supplier_db = prepare_supplier_database(path=SUPPLIER_DB_PATH)

    if supplier_db.empty:
        st.warning("Upload DATABASE SUP.xlsx dulu supaya supplier bisa otomatis terisi.")
    else:
        st.success(f"Database supplier terbaca: {len(supplier_db)} item.")

    st.header("Status Data")

    if st.session_state.all_data:
        status_df = pd.DataFrame(st.session_state.all_data)
        counts = status_df.groupby("Unit").size().reindex(UNITS, fill_value=0)
    else:
        counts = pd.Series(0, index=UNITS)

    for unit_name, total in counts.items():
        st.write(f"**{unit_name}**: {total} baris")

    if st.session_state.all_data and st.button("Reset Data Sementara"):
        st.session_state.all_data = []
        st.session_state.excel_file = None
        st.rerun()

unit = st.selectbox("Pilih Unit Upload", UNITS)

uploaded_files = st.file_uploader(
    f"Upload PDF untuk {unit}",
    type=["pdf"],
    accept_multiple_files=True,
    key=unit,
)

col_add, col_hint = st.columns([1, 2])
with col_add:
    add_clicked = st.button(
        f"Tambah Data {unit}",
        disabled=not uploaded_files,
        use_container_width=True,
    )

with col_hint:
    if uploaded_files:
        st.info(f"{len(uploaded_files)} file PDF siap ditambahkan untuk {unit}.")
    else:
        st.info("Pilih PDF purchase order untuk unit yang sedang dipilih.")

if add_clicked:
    added_rows = 0
    empty_files = []

    with st.spinner(f"Membaca PDF untuk {unit}..."):
        for uploaded_file in uploaded_files:
            rows = parse_pdf(uploaded_file, unit, supplier_db)
            if rows:
                st.session_state.all_data.extend(rows)
                added_rows += len(rows)
            else:
                empty_files.append(uploaded_file.name)

    st.session_state.excel_file = None

    if added_rows:
        st.success(f"Data {unit} berhasil ditambahkan: {added_rows} baris.")
    if empty_files:
        st.warning(
            "Tidak ada baris item yang cocok dari file: "
            + ", ".join(empty_files)
            + ". Cek format PDF atau teks hasil extract."
        )

if st.session_state.all_data:
    df_preview = pd.DataFrame(st.session_state.all_data)
    preview_columns = [column for column in COLUMNS if column in df_preview.columns]

    st.divider()
    st.subheader("Preview Data Sementara")
    st.dataframe(df_preview[preview_columns], use_container_width=True, hide_index=True)

    if st.button("Convert Semua Unit", use_container_width=True):
        st.session_state.excel_file = build_excel(st.session_state.all_data)
        st.success("Excel berhasil dibuat: 1 file dengan sheet per unit.")

    if st.session_state.excel_file:
        st.download_button(
            "Download Excel",
            data=st.session_state.excel_file,
            file_name="hasil_final.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
else:
    st.divider()
    st.info("Belum ada data. Upload PDF untuk salah satu unit, lalu klik Tambah Data.")
